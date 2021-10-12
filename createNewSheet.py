#Module Imports
from robin_stocks import robinhood
import pyotp
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font, Alignment, Border,Side, Color
from openpyxl.formatting.rule import ColorScaleRule, FormatObject
import datetime
import logging
import os

#File Imports
import config

logging.basicConfig(filename='debug.log', level=logging.INFO, format='%(message)s')

boldFont = Font(bold=True)
centerAlignment = Alignment(horizontal='center',vertical='center')
allBorders = Border(left=Side(style="thin", color="000000"),
					right=Side(style="thin", color="000000"),
					top=Side(style="thin", color="000000"),
					bottom=Side(style="thin", color="000000"))
currencyFormat = r'_("$"* #,##0.00_)_("$"* \(#,##0.00\)_("$"* "-"??_)_(@_)'
percentFormat = '0.00%'
threeColorScaleFormatting = ColorScaleRule(start_type="min",start_color='F8696B',mid_type="num",mid_value=0,mid_color="FFFFFF",end_type="max",end_color='63BE7B')

def main():
	stocks = getStocks()
	createSheet(stocks)

def getStocks():
	totp = pyotp.TOTP(config.get_robinhood_mfa_code()).now()
	login = robinhood.login(config.get_robinhood_username(),config.get_robinhood_password(),mfa_code=totp)

	my_stocks = robinhood.account.build_holdings()

	robinhood.authentication.logout()
	return my_stocks

def createSheet(stocks):
	sheetName = str(datetime.datetime.today().year)
	if not (os.path.exists(config.get_excel_fileName())):
		wb = Workbook()
		sheet = wb.active
		sheet.title = sheetName
	else:
		wb = load_workbook(filename=config.get_excel_fileName())	
		sheet = wb.create_sheet(sheetName,0)

	#First Section: Amount Invested
	columnToStart = 1
	endOfSection = createSection(sheet,columnToStart,stocks,"Amount Invested",sheetName)

	#Second Section: Equity
	endOfSection = createSection(sheet,endOfSection+2,stocks,"Equity",sheetName)
	#Third Section: Total Return
	endOfSection = createSection(sheet,endOfSection+2,stocks,"Total Return",sheetName)
	#Fourth Section: Total Percent Return
	endOfSection = createSection(sheet,endOfSection+2,stocks,"Total Percent Return",sheetName)
	#Fifth Section: Day Over Day Return
	endOfSection = createSection(sheet,endOfSection+2,stocks,"Day Over Day Return",sheetName)
	#Sixth Section: Day Over Day Percent Change
	createSection(sheet,endOfSection+2,stocks,"Day Over Day Percent Change",sheetName)


	wb.save(config.get_excel_fileName())

#Creates the section with dates and returns the index of the most right column
def createSection(sheet,columnToStart,stocks, sectionName,year):
	numberOfStocks = len(stocks)

	#Merged Top Bar
	updateCell(sheet,1,columnToStart,sectionName,boldFont,centerAlignment,allBorders,None)
	
	sheet.merge_cells(start_row=1,start_column=columnToStart,end_row=1,end_column=columnToStart+numberOfStocks+1)

	#Column Headers

	#Date Column
	updateCell(sheet,2,columnToStart,"Date",boldFont,centerAlignment,allBorders,None)

	#Stock Columns
	columnToInsert = columnToStart + 1
	for stock in stocks:
		updateCell(sheet,2,columnToInsert,stock,boldFont,centerAlignment,allBorders,None)
		columnToInsert = columnToInsert + 1
	
	#Total Column
	updateCell(sheet,2,columnToInsert,"Total",boldFont,centerAlignment,allBorders,None)

	#Dates
	yearEnd = datetime.datetime(int(year),12,31)
	dateToInsert = datetime.datetime(int(year),1,1)
	rowToInsert = 3
	while dateToInsert <=yearEnd:
		if dateToInsert.weekday() != 5 and dateToInsert.weekday() !=6:
			updateCell(sheet,rowToInsert,columnToStart,dateToInsert.date(),None,None,allBorders,None)
			addFormulas(sheet, sectionName, columnToStart,columnToInsert, rowToInsert,stocks)

			rowToInsert = rowToInsert+1
		dateToInsert = dateToInsert + datetime.timedelta(days=1)


	return columnToInsert

def addFormulas(sheet, sectionName, leftBoundry, rightBoundry, rowToInsert,stocks):
	amountInvestedStock = 2
	#Start at first Stock (skipping date column), move over by the number of stocks, and then three more
	equityStock = amountInvestedStock + len(stocks) + 3
	if sectionName == "Amount Invested" or sectionName == "Equity":
		updateCell(sheet,rowToInsert,rightBoundry,f"=SUM({numberToColumn(leftBoundry+1)}{rowToInsert}:{numberToColumn(rightBoundry-1)}{rowToInsert})",None,None,allBorders,currencyFormat)
		return
	elif sectionName == "Total Return":
		for column in range(leftBoundry+1,rightBoundry+1):
			updateCell(sheet,rowToInsert,column,f"={numberToColumn(equityStock)}{rowToInsert} - {numberToColumn(amountInvestedStock)}{rowToInsert}",None,None,allBorders,currencyFormat)
			amountInvestedStock = amountInvestedStock + 1
			equityStock = equityStock +1
		
	elif sectionName == "Total Percent Return":
		#Skips the first row for both percent changes
		if rowToInsert == 3:
			return
		for column in range(leftBoundry+1,rightBoundry+1):
			equityStockColumn = numberToColumn(equityStock)
			amountInvestedStockColumn = numberToColumn(amountInvestedStock)

			#(Equity - Invested) / Invested
			formula = f"=({equityStockColumn}{rowToInsert}-{amountInvestedStockColumn}{rowToInsert})/ABS({amountInvestedStockColumn}{rowToInsert})"
			updateCell(sheet,rowToInsert,column,formula,None,None,allBorders,percentFormat)
			amountInvestedStock = amountInvestedStock + 1
			equityStock = equityStock +1
		
	elif sectionName == "Day Over Day Return":
		#Skips the first row for both percent changes
		if rowToInsert == 3:
			return
		for column in range(leftBoundry+1,rightBoundry+1):
			equityStockColumn = numberToColumn(equityStock)
			amountInvestedStockColumn = numberToColumn(amountInvestedStock)

			#(Equity - Invested) / Invested
			formula = f"={equityStockColumn}{rowToInsert}-{equityStockColumn}{rowToInsert-1}"
			updateCell(sheet,rowToInsert,column,formula,None,None,allBorders,currencyFormat)
			amountInvestedStock = amountInvestedStock + 1
			equityStock = equityStock +1
	elif sectionName == "Day Over Day Percent Change":
		for column in range(leftBoundry+1,rightBoundry+1):
			equityStockColumn = numberToColumn(equityStock)
			amountInvestedStockColumn = numberToColumn(amountInvestedStock)

			#(Equity - Invested) / Invested
			formula = f"=({equityStockColumn}{rowToInsert}-{equityStockColumn}{rowToInsert-1})/ABS({equityStockColumn}{rowToInsert-1})"
			updateCell(sheet,rowToInsert,column,formula,None,None,allBorders,percentFormat)
			amountInvestedStock = amountInvestedStock + 1
			equityStock = equityStock +1
	#Conditional Formatting
	sheet.conditional_formatting.add(f'{numberToColumn(leftBoundry+1)}{rowToInsert}:{numberToColumn(column-1)}{rowToInsert}',threeColorScaleFormatting)

#Found this solution here: https://stackoverflow.com/questions/23861680/convert-spreadsheet-number-to-column-letter
def numberToColumn(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


def updateCell(sheet,row,column,value,font,alignment,border,number_format):
	cellToChange = sheet.cell(row,column)

	if number_format != None:
		cellToChange.number_format = number_format

	cellToChange.font = font
	cellToChange.alignment = alignment
	cellToChange.border = border
	cellToChange.value = value
if __name__ == "__main__":
	main()