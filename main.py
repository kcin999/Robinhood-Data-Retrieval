from robin_stocks import robinhood
import config
import pyotp
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border,Side
import datetime
import logging
import sys
import traceback

logging.basicConfig(filename='debug.log', level=logging.INFO, format='%(message)s')

boldFont = Font(bold=True)
centerAlignment = Alignment(horizontal='center',vertical='center')
allBorders = Border(left=Side(style="thin", color="000000"),
					right=Side(style="thin", color="000000"),
					top=Side(style="thin", color="000000"),
					bottom=Side(style="thin", color="000000"))
currencyFormat = r'_("$"* #,##0.00_)_("$"* \(#,##0.00\)_("$"* "-"??_)_(@_)'

def main():
	stocks = getStocks()
	loadIntoExcel(stocks)

#Grabs all of the stock data that is currently in my portfolio
def getStocks():
	totp = pyotp.TOTP(config.get_robinhood_mfa_code()).now()
	login = robinhood.login(config.get_robinhood_username(),config.get_robinhood_password(),mfa_code=totp)

	my_stocks = robinhood.account.build_holdings()

	robinhood.authentication.logout()
	return my_stocks

def loadIntoExcel(stocks):
	workbook = load_workbook(filename=config.get_excel_fileName())	
	yearNumber = str(datetime.datetime.today().year)
	sheet = workbook[yearNumber]

	dateColumn,dateRow = findCellCoordinates(sheet,"Date",2)
	totalColumn,totalRow = findCellCoordinates(sheet,"Total",2)

	firstDateColumn, firstDateRow = findCellCoordinates(sheet,"Date",1)

	#Get Column Headers
	stockHeaders = []
	for column in sheet.iter_cols(min_row=2,max_row=2,min_col=dateColumn+1,max_col=totalColumn-1):
		for cell in column:
			stockHeaders.append(cell.value)

	#Get row that needs to be updated. 
	rowToUpdate = 0
	for column in sheet.iter_cols(min_row=3,min_col=dateColumn,max_col=dateColumn):
		todayValue = datetime.date.today()
		for cell in column:
			cellValue = cell.value.date()
			if(cellValue == todayValue):
				rowToUpdate = cell.row
				break
			elif (cell.value.date() >= todayValue):
				todayString = todayValue.strftime("%m/%d/%Y")
				logging.info(f"Cannot find date {todayString} in table.  Maybe it's the weekend?")
				rowToUpdate = -1
				break
		break


	#Update row
	#See above. rowToUpdate will be -1 if date is not found in table.
	if rowToUpdate == -1:
		return
	else:
		for i in range(0,len(stockHeaders)):
			#If stock header cannot be found in equity, then value becomes 0
			try:
				equity = float(stocks[stockHeaders[i]]['equity'])
			except:
				equity = 0
			try:
				amountInvested = float(stocks[stockHeaders[i]]['equity']) - float(stocks[stockHeaders[i]]['equity_change'])
			except:
				amountInvested = 0

			updateCell(sheet,rowToUpdate,dateColumn+1+i,equity,None,None,allBorders,currencyFormat)
			updateCell(sheet,rowToUpdate,firstDateColumn + 1+i,amountInvested,None,None,allBorders,currencyFormat)

		workbook.save(config.get_excel_fileName())

#Finds the cellValue within the sheet.  
#Checks the second row for the value
#Can find the occurance of the value. For example, if you want to find the second time "foundME" appears in sheet, you would call findCellCorridnates(sheet,"foundME",2) 
def findCellCoordinates(sheet,cellValue, occurance):
	count = 0
	for column in sheet.iter_cols(min_row=2,min_col=1,max_row=2):
		for cell in column:
			if cell.value == cellValue:
				count = count + 1
				if (count == occurance):
					return (cell.column,cell.row)


def updateCell(sheet,row,column,value,font,alignment,border,number_format):
	cellToChange = sheet.cell(row,column)

	if number_format != None:
		cellToChange.number_format = number_format

	cellToChange.font = font
	cellToChange.alignment = alignment
	cellToChange.border = border
	cellToChange.value = value

if __name__ == "__main__":
	try:
		main()
	except Exception as e:
		exc_type, exc_obj, exc_tb = sys.exc_info()
        logging.info("\t\t" + str(e) + " at " + datetime.datetime.now().strftime("%m/%d/%Y %H:%M:%S:%f") )
        logging.info("Call Stack:")
        callStack = traceback.format_exc()  
        logging.info( callStack + "\n\n")