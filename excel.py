#Module Imports
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font, Alignment, Border,Side
from openpyxl.formatting.rule import ColorScaleRule
import datetime
import os

#File Imports
import config

#Global Variables
sectionHeaders = ["Amount Invested", "Equity", "Total Return", "Total Percent Return", "Day Over Day Return", "Day Over Day Percent Change"]

boldFont = Font(bold=True)
centerAlignment = Alignment(horizontal='center',vertical='center')
allBorders = Border(left=Side(style="thin", color="000000"),
					right=Side(style="thin", color="000000"),
					top=Side(style="thin", color="000000"),
					bottom=Side(style="thin", color="000000"))
currencyFormat = r'_("$"* #,##0.00_)_("$"* \(#,##0.00\)_("$"* "-"??_)_(@_)'
percentFormat = '0.00%'
threeColorScaleFormatting = ColorScaleRule(start_type="min",start_color='F8696B',mid_type="num",mid_value=0,mid_color="FFFFFF",end_type="max",end_color='63BE7B')
row = 1
column = 1
sheetName = str(datetime.datetime.today().year)

def createWorkbook():
	wb = Workbook()

	wb.save(config.get_excel_fileName())
	return wb

def createSheet(workbook, stocks):
	sheet = workbook.create_sheet(sheetName,0)

	createAllSections(sheet,stocks)

	workbook.save(config.get_excel_fileName())

def createAllSections(sheet, stocks):	
	for section in sectionHeaders:
		createSection(sheet,section, stocks)

def createSection(sheet, sectionName, stocks):
	global row, column
	numberOfStocks = len(stocks)

	writeMergedHeader(sheet, sectionName, numberOfStocks)
	row += 1

	writeStockHeaders(sheet, sectionName, stocks)

	column += 2
	row = 1

def writeMergedHeader(sheet, sectionName, numberOfStocks):
	global row
	row = 1
	#Writes Amount Invested Merged Cell
	updateCell(sheet, sectionName,boldFont,centerAlignment,allBorders,None)
	sheet.merge_cells(start_row=row, start_column=column,end_row=1,end_column= column +numberOfStocks+1 )

def writeStockHeaders(sheet, sectionName, stocks):
	global column
	#Write Column Headers such as Date, Stock, Total
	updateCell(sheet, "Date", boldFont, centerAlignment, allBorders, None)
	column += 1

	for stock in stocks:
		updateCell(sheet, stock, boldFont, centerAlignment, allBorders, None)
		column +=1

	updateCell(sheet,"Total",boldFont,centerAlignment,allBorders,None)

def getWorkbook(stocks):
	if not (os.path.exists(config.get_excel_fileName())):
		wb = createWorkbook()
	else:
		wb = load_workbook(config.get_excel_fileName())

	
	if sheetName not in wb.sheetnames:
		createSheet(wb, stocks)
	
	return load_workbook(filename=config.get_excel_fileName())

def addDataToExcel(stocks):
	global column
	

	workbook = getWorkbook(stocks)	

	sheet = workbook[sheetName]

	stocksInSheet = getStocksInSheet(sheet)

	#reset's column
	column = 1	
	addNewStocksToSheet(sheet,stocksInSheet, stocks)

	findRowToAddData(sheet)

	stocksInSheet = getStocksInSheet(sheet)

	addAmountInvested(sheet, stocks, stocksInSheet)
	
	addEquity(sheet, stocks, stocksInSheet)

	addFormulas(sheet,stocksInSheet)

	workbook.save(config.get_excel_fileName())

def findRowToAddData(sheet):
	global row

	for rowInExcel in sheet.iter_rows(min_row=row):
		if rowInExcel[0].value == datetime.datetime.today().replace(hour=0,minute=0,second=0,microsecond=0):
			return
		row += 1

def addNewStocksToSheet(sheet,stocksInSheet, stocks):
	global row, column
	"""If there are any new stocks in the robinhood account, that have not been in the spreadsheet, this will add them
	
	If there are no new stocks, this will just continue return
	"""
	missingStocks = findMissingStocks(stocksInSheet, stocks)

	for stock in missingStocks:
		addStockToSheet(sheet, stocksInSheet,stock)

	#Resets the row to 1 in order to have the correct cells merged
	row = 1
	column = 1
	#Unmerges Already Merged Sections
	mergedCells = []
	for cellgroup in sheet.merged_cells.ranges:
		mergedCells.append(str(cellgroup))

	for range in mergedCells:
		sheet.merged_cells.remove(range)

	#Merges new section
	for section in sectionHeaders:
		writeMergedHeader(sheet,section, len(stocks))
		column += 2 + len(stocks) + 1

def findMissingStocks(stocksInSheet, stocks):
	stocksAlreadyAdded = []

	#Finds duplicates
	for stock in stocks:
		if stock in stocksInSheet:
			stocksAlreadyAdded.append(stock)

	newStocks = stocks.copy()
	
	for stock in stocksAlreadyAdded:
		del newStocks[stock]

	return newStocks

def addStockToSheet(sheet, stocksInSheet, stock):
	global column
	global row

	row = 2

	#Amount Invested
	column = len(stocksInSheet) + column + 1
	sheet.insert_cols(column)
	updateCell(sheet,stock,boldFont,centerAlignment,allBorders,None)


	#Equity
	column = len(stocksInSheet) + column + 4
	sheet.insert_cols(column)
	updateCell(sheet,stock,boldFont,centerAlignment,allBorders,None)



	#Equity
	column = len(stocksInSheet) + column + 4
	sheet.insert_cols(column)
	updateCell(sheet,stock,boldFont,centerAlignment,allBorders,None)


	#Total Return
	column = len(stocksInSheet) + column + 4
	sheet.insert_cols(column)
	updateCell(sheet,stock,boldFont,centerAlignment,allBorders,None)


	#Total Percent Return
	column = len(stocksInSheet) + column + 4
	sheet.insert_cols(column)
	updateCell(sheet,stock,boldFont,centerAlignment,allBorders,None)


	#Day Over Day Return
	column = len(stocksInSheet) + column + 4
	sheet.insert_cols(column)
	updateCell(sheet,stock,boldFont,centerAlignment,allBorders,None)

	
	#Day Over Day Percent Change
	column = len(stocksInSheet) + column + 4
	sheet.insert_cols(column)
	updateCell(sheet,stock,boldFont,centerAlignment,allBorders,None)

#Gets the list of stocks that are already in the sheet
def getStocksInSheet(sheet):
	rowOfHeaders = 2
	stockList = []
	columnToCheck = 1

	while columnToCheck != -1:
		cellValue = sheet.cell(rowOfHeaders,columnToCheck).value
		columnToCheck += 1
		if cellValue == "Date":
			pass
		elif cellValue == "Total":
			columnToCheck = -1
		else:
			stockList.append(cellValue)

	return stockList

def addAmountInvested(sheet, stocks, stocksInSheet):
	global column, row

	column = 1
	dateColumnIndex = 1

	#Adds date
	updateCell(sheet, datetime.date.today(),None,None,allBorders,None)

	#Adds Amount Invested	
	for stock in stocks:
		column = dateColumnIndex + stocksInSheet.index(stock) + 1

		try:
			amountInvested = float(stocks[stock]['equity']) - float(stocks[stock]['equity_change'])
		except:
			amountInvested = 0
		updateCell(sheet, amountInvested, None, None, allBorders, currencyFormat)

	#Writes Summation for total column
	column += 1
	updateCell(sheet,f"=SUM({numberToColumn(dateColumnIndex + 1)}{row}:{numberToColumn(dateColumnIndex + len(stocksInSheet))}{row})",None,None,allBorders,currencyFormat)

def addEquity(sheet,stocks,stocksInSheet):
	global column

	#Column Offset for the next category
	column += 2 
	dateColumnIndex = column

	#Adds date
	updateCell(sheet, datetime.date.today(),None,None,allBorders,None)

	#Adds Amount Invested	
	for stock in stocks:
		column = dateColumnIndex + stocksInSheet.index(stock) + 1

		try:
			equity = float(stocks[stock]['equity'])
		except:
			equity = 0

		updateCell(sheet, equity, None, None, allBorders, currencyFormat)

	#Writes Summation for total column
	column += 1
	updateCell(sheet,f"=SUM({numberToColumn(dateColumnIndex + 1)}{row}:{numberToColumn(dateColumnIndex + len(stocksInSheet))}{row})",None,None,allBorders,currencyFormat)

def addFormulas(sheet,stocksInSheet):
	addTotalReturnFormulas(sheet,stocksInSheet)
	addTotalPercentReturnFormulas(sheet,stocksInSheet)
	addDayOverDayReturnFormulas(sheet,stocksInSheet)
	addDayOverDayPercentReturnFormulas(sheet,stocksInSheet)

def addTotalReturnFormulas(sheet,stocksInSheet):
	global column
	column += 2
	dateColumnIndex = column

	updateCell(sheet, datetime.date.today(),None,None,allBorders,None)

	#Plus 1 is to grab the total column
	for i in range(0, len(stocksInSheet)+1):
		column = dateColumnIndex + i + 1

		amountInvestedColumn = numberToColumn(i+ 2)
		equityColumn = numberToColumn(len(stocksInSheet) + i +5)
		formula = f"={equityColumn}{row} - {amountInvestedColumn}{row}"
		
		updateCell(sheet, formula, None,None, allBorders, currencyFormat)
		
	#Adds three color scale formatting
	sheet.conditional_formatting.add(f'{numberToColumn(dateColumnIndex + 1)}{row}:{numberToColumn(column-1)}{row}',threeColorScaleFormatting)

def addTotalPercentReturnFormulas(sheet,stocksInSheet):
	global column
	column += 2
	dateColumnIndex = column

	updateCell(sheet, datetime.date.today(),None,None,allBorders,None)

	#Plus 1 is to grab the total column
	for i in range(0, len(stocksInSheet)+1):
		column = dateColumnIndex + i + 1

		amountInvestedColumn = numberToColumn(i+ 2)
		equityColumn = numberToColumn(len(stocksInSheet) + i +5)
		formula = f"=({equityColumn}{row} - {amountInvestedColumn}{row})/ABS({amountInvestedColumn}{row})"
		
		updateCell(sheet, formula, None,None, allBorders, percentFormat)

	#Adds three color scale formatting
	sheet.conditional_formatting.add(f'{numberToColumn(dateColumnIndex + 1)}{row}:{numberToColumn(column-1)}{row}',threeColorScaleFormatting)

def addDayOverDayReturnFormulas(sheet,stocksInSheet):
	global column
	column += 2
	dateColumnIndex = column

	updateCell(sheet, datetime.date.today(),None,None,allBorders,None)

	#Plus 1 is to grab the total column
	for i in range(0, len(stocksInSheet)+1):
		column = dateColumnIndex + i + 1

		equityColumn = numberToColumn(len(stocksInSheet) + i +5)
		formula = f"={equityColumn}{row} - {equityColumn}{row-1}"
		
		updateCell(sheet, formula, None,None, allBorders, currencyFormat)

	#Adds three color scale formatting
	sheet.conditional_formatting.add(f'{numberToColumn(dateColumnIndex + 1)}{row}:{numberToColumn(column-1)}{row}',threeColorScaleFormatting)

def addDayOverDayPercentReturnFormulas(sheet,stocksInSheet):
	global column
	column += 2
	dateColumnIndex = column

	updateCell(sheet, datetime.date.today(),None,None,allBorders,None)

	#Plus 1 is to grab the total column
	for i in range(0, len(stocksInSheet)+1):
		column = dateColumnIndex + i + 1

		equityColumn = numberToColumn(len(stocksInSheet) + i +5)
		formula = f"=({equityColumn}{row} - {equityColumn}{row-1})/ABS({equityColumn}{row-1})"
		
		updateCell(sheet, formula, None,None, allBorders, percentFormat)

	#Adds three color scale formatting
	sheet.conditional_formatting.add(f'{numberToColumn(dateColumnIndex + 1)}{row}:{numberToColumn(column-1)}{row}',threeColorScaleFormatting)

def updateCell(sheet,value,font,alignment,border,number_format):
	cellToChange = sheet.cell(row,column)

	if number_format != None:
		cellToChange.number_format = number_format

	cellToChange.font = font
	cellToChange.alignment = alignment
	cellToChange.border = border
	cellToChange.value = value

#Found this solution here: https://stackoverflow.com/questions/23861680/convert-spreadsheet-number-to-column-letter
def numberToColumn(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string